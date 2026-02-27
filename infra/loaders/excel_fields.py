from __future__ import annotations

from pathlib import Path
import openpyxl

EXCEL_FILE = "Campos_tabela_de_MO_otimizado.xlsx"
SHEET_NAME = "Schema_MO"


def _candidates() -> list[Path]:
    here = Path(__file__).resolve()
    repo_root = here.parents[2]        # /workspaces/sisroad-dash
    workspace_root = repo_root.parent  # /workspaces
    return [repo_root / EXCEL_FILE, workspace_root / EXCEL_FILE, Path.cwd() / EXCEL_FILE]


def get_excel_path() -> Path:
    for p in _candidates():
        if p.exists():
            return p
    tried = "\n".join(str(p) for p in _candidates())
    raise FileNotFoundError(f"Excel não encontrado. Tentei:\n{tried}")


def get_excel_schema() -> list[dict]:
    """
    Lê a aba 'Schema_MO' e retorna:
      [{"name": "...", "sql": "..."}]

    Usa o tipo SQL exatamente como está no Excel.
    """
    p = get_excel_path()
    wb = openpyxl.load_workbook(p, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Aba '{SHEET_NAME}' não encontrada. Abas: {wb.sheetnames}")

    ws = wb[SHEET_NAME]

    # linha 1 é cabeçalho: Nome | Tipo de Dado
    schema: list[dict] = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        sql = ws.cell(r, 2).value

        if not name:
            continue

        name = str(name).strip()
        sql = str(sql).strip() if sql else "NVARCHAR(255)"

        schema.append({"name": name, "sql": sql})

    if not schema:
        raise ValueError("Schema vazio. Verifique a aba Schema_MO (Nome/Tipo de Dado).")

    return schema


def input_type_from_sql(sql_type: str) -> str:
    """
    Define o tipo do input do formulário baseado no SQL do Excel.
    (sem 'mapear excel': usamos o SQL que já está no Excel)
    """
    t = (sql_type or "").upper()

    # Datas
    if "DATE" in t:          # DATE, DATETIME, DATETIME2...
        return "date" if "TIME" not in t else "text"

    # Números
    numeric_tokens = ("INT", "DECIMAL", "NUMERIC", "FLOAT", "REAL", "MONEY")
    if any(tok in t for tok in numeric_tokens):
        return "number"

    # Boolean (BIT)
    if "BIT" in t:
        return "text"  # depois você pode evoluir para switch

    # Default texto
    return "text"